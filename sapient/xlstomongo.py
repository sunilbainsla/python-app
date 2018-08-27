import csv
import pymongo
from openpyxl import load_workbook
wb = load_workbook(filename='D:\MIP\Manchester-housing.xlsx')
ws = wb['Sheet1']
try:
    myclient = pymongo.MongoClient("mongodb://localhost:27017/")
    mydb = myclient["mydatabase"]
    mycol = mydb["MIP_tab"]
    for i, row in enumerate(ws.iter_rows()):
        if i == 0:
            continue
        else:
            for cell in row:
                print(cell.value, end=" ")
except mydb.error as e:
    mycol.rollback()
    raise
finally:
    myclient.close()
    wb.close()